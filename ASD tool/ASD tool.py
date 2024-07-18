import numpy as np
import numpy.matlib
import tensorflow as tf
from scipy import signal
from scipy.signal import butter, lfilter, freqz
from scipy.signal import find_peaks
from scipy.signal import argrelextrema
import pandas as pd
import xlrd
import os
import librosa
import librosa.display
import math
import sys
import sk_dsp_comm.sigsys as ss
import sk_dsp_comm.fir_design_helper as fir_d
import openpyxl
from openpyxl import Workbook

"""# **load data**"""

# Commented out IPython magic to ensure Python compatibility.
# read the data table:

xlrd.xlsx.ensure_elementtree_imported(False, None)
xlrd.xlsx.Element_has_iter = True
data_table = xlrd.open_workbook('C:/Users/77sha/Desktop/final project/New Data For Syl Segmentation_4.xlsx').sheet_by_index(0)
mother = data_table.col_values(0, 1)
matgen = data_table.col_values(1, 1)
name = data_table.col_values(2, 1)
sex = data_table.col_values(3, 1)
pupgen = data_table.col_values(4, 1)
age = data_table.col_values(5, 1)
session = data_table.col_values(6, 1)
rec_num = data_table.col_values(7, 1)

#find path and load recordings:
SignalVec = []
signal_name = []
sr = 250000
for i in range(len(mother)):
  path = 'C:/Users/77sha/Desktop/USV Recording/2022/{}_{}/{}_{}/day_{}/session{}/{}.wav'.format(mother[i], matgen[i], name[i], pupgen[i], int(age[i]), int(session[i]), rec_num[i]) #find path of each recording
  if not os.path.exists('{}'.format(path)):
    path = 'C:/Users/77sha/Desktop/USV Recording/2022/{}_{}/{}_{}/day_{}/session{}/{}.WAV'.format(mother[i], matgen[i], name[i], pupgen[i], int(age[i]), int(session[i]), rec_num[i])
    if not os.path.exists('{}'.format(path)):
      print(i)
      continue
  signal_name.append(path)
  rec, rate = librosa.load(path, sr) #opens recordings and sample rate
  SignalVec.append(rec)

"""# **Segmentation**"""

from Segmentation import *

Fs = rate
FrameLength = 0.006
Overlap = 0.7
thresh = 20
harmony_th = 0.009

siz = len(SignalVec)
book = Workbook()
sheet = book.active
title = ['Path','Mother','Mother Genotype','Name','Sex','Offspring Genotype','Day','Session','Recording Number','Start point(s)','End point(s)','Duration (time)']
sheet.append(title)
for s2 in range(siz):
  signal = SignalVec[s2]
  signal = Preprocessing(signal,Fs)
  # if there is a 'silent' start (zeros), skipping to the "real" start:
  ind = np.where(signal == 0)
  is_empty = ind[0].size == 0
  if not(is_empty) and ind[0][0] == 0:
   DiffInd = np.diff(np.diff(ind))
   ind2 = np.where(DiffInd != 0)
   is_empty = ind2[0].size == 0
   if not(is_empty):
     for i in range(0,len(signal)-int(ind2[0])):
       signal[i] = signal[i+int(ind2[0])]
     i = range(len(signal)-int(ind2[0]),len(signal))
     signal = np.delete(signal,i)
   else:
     ind2 = [[0],[0]]
  else:
    ind2 = [[0],[0]]

  _,_,_,_,ClassLPC,SyllabelVec,SignalPath = Syllables_Detection2(signal,Fs,FrameLength,Overlap, thresh, harmony_th, signal_name[s2], ind2)

  if any(SyllabelVec):
    StartEndNew = Rearrange_signal(signal,Fs,ClassLPC.time1) #StartEndNew - times vector
    StEndMatF = Check_length_Call(StartEndNew)
    print(StEndMatF)

    
    for i in range(len(StEndMatF)):
      Duration = StEndMatF[i][1] - StEndMatF[i][0]
      new_row = [signal_name[s2],mother[s2],matgen[s2],name[s2],sex[s2],pupgen[s2],age[s2],session[s2],rec_num[s2],StEndMatF[i][0],StEndMatF[i][1],Duration]
      sheet.append(new_row)
book.save('C:/Users/77sha/Desktop/MatPython/New Segmentatio Data For Final Class_4.xlsx')


"""# **Features: ISI and StartEnFreq**"""


data_table = xlrd.open_workbook('C:/Users/77sha/Desktop/MatPython/New Segmentatio Data For Final Class_4.xlsx').sheet_by_index(0)
motherSyl = data_table.col_values(1, 1)
matgenSyl = data_table.col_values(2, 1)
nameSyl = data_table.col_values(3, 1)
sexSyl = data_table.col_values(4, 1)
pupgenSyl = data_table.col_values(5, 1)
ageSyl = data_table.col_values(6, 1)
sessionSyl = data_table.col_values(7, 1)
rec_numSyl = data_table.col_values(8, 1)
startSyl = data_table.col_values(9, 1)
endSyl = data_table.col_values(10, 1)

from Features import *

ISI = ISI_time(rec_numSyl,startSyl,endSyl)
startF,endF = StartEndFreq(SignalVec,siz,mother,name,age,session,rec_num,motherSyl,nameSyl,ageSyl,sessionSyl,rec_numSyl,startSyl,endSyl,rate)

y = 2
workbook = openpyxl.load_workbook('C:/Users/77sha/Desktop/MatPython/New Segmentatio Data For Final Class_4.xlsx')
worksheet = workbook.worksheets[0]
worksheet.insert_cols(13,15)
cell_title1 = worksheet.cell(row=1, column=13)
cell_title1.value = 'ISI_time'
cell_title2 = worksheet.cell(row=1, column=14)
cell_title2.value = 'Start Point (Hz)'
cell_title3 = worksheet.cell(row=1, column=15)
cell_title3.value = 'End Point (Hz)'
for x in range(len(ISI)):
    cell_to_write = worksheet.cell(row=y, column=13)
    cell_to_write.value = ISI[x]
    cell_to_write = worksheet.cell(row=y, column=14)
    cell_to_write.value = startF[x]
    cell_to_write = worksheet.cell(row=y, column=15)
    cell_to_write.value = endF[x]
    y += 1
workbook.save('C:/Users/77sha/Desktop/MatPython/New Segmentatio Data For Final Class_4.xlsx')


"""# **Syllabel classification**"""

from statistics_generator import *

model_path = 'C:/Users/77sha/Desktop/Python/model_weights.h6'
model = keras.models.load_model(model_path, custom_objects={'KerasLayer':hub.KerasLayer})

samples = Syl_Class_Vec(model,ageSyl,matgenSyl,pupgenSyl,motherSyl,nameSyl,sexSyl,sessionSyl,rec_numSyl,startSyl,endSyl)
print(samples)
np.save('C:/Users/77sha/Desktop/MatPython/model_prediction_recordings_4.npy', samples)

syl_num = []
for i in range(len(samples)):
  for j in range(len(samples[i].syls)):
    if np.max(samples[i].syls[j])<0.5:
      temp = 10
    else:
      temp = np.argmax(samples[i].syls[j])
    samples[i].syls[j] = []
    samples[i].syls[j] = temp
    syl_num.append(samples[i].syls[j])
    print(samples[i].syls[j])


y = 2
workbook = openpyxl.load_workbook('C:/Users/77sha/Desktop/MatPython/New Segmentatio Data For Final Class_4.xlsx')
worksheet = workbook.worksheets[0]
worksheet.insert_cols(16)
cell_title = worksheet.cell(row=1, column=16)
cell_title.value = 'Syllable number'
for x in range(len(syl_num)):
    cell_to_write = worksheet.cell(row=y, column=16)
    cell_to_write.value = syl_num[x]
    y += 1
workbook.save('C:/Users/77sha/Desktop/MatPython/New Segmentatio Data For Final Class_4.xlsx')

'''#Features extracting table#'''

from audio_feature_extraction_reduction_by_recording import *

dataset = pd.read_excel('C:/Users/77sha/Desktop/MatPython/New Segmentatio Data For Final Class_4.xlsx')

# extract only the relevant columns / features
X = dataset[["Name", "Day", "Session", "Start Point (Hz)", "End Point (Hz)", "Duration (time)", "Syllable number", "Recording Number", "Mother Genotype", "Sex", "ISI_time", "Offspring Genotype"]]

mouse_final_data = feature_extraction(X)
# export data to csv file for further use - COLAB
np.savetxt("C:/Users/77sha/Desktop/MatPython/processed_data_for_final_classification_REDUCTION_BY_RECORDING_4.csv", X=mouse_final_data, delimiter=",")


